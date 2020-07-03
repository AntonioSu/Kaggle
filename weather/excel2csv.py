import openpyxl

#将Excel转化为csv格式
def Write(path):
    # 选择工作表
    wb = openpyxl.load_workbook(path)
    sheet = wb['weather']
    path = 'data.csv'
    with open(path, 'w') as f:
        row=1
        while True:
            if sheet.cell(row=row, column=1).value == None:
                break
            data=[]
            for column in range(1,13):
                data.append(str(sheet.cell(row=row, column=column).value))
            data=','.join(data)
            f.write(data+'\n')
            row+=1
        f.flush()

if __name__=='__main__':
    # path='merge.xlsx'
    # get_initial_data(path)
    path = 'data.xlsx'
    Write(path)
    print('done')